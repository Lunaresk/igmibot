from ..bottoken import getConn
from openpyxl import (Workbook, load_workbook)
from datetime import date

# Arrays for openpyxl
months = ["Januar", "Februar", "März", "April", "Mai", "Juni", "Juli", "August", "September", "Oktober", "November", "Dezember"]
heads = ["Vorname", "Nachname", "Email-Adresse", "Kaffee/Tee", "Softgetränke", "Wasser", "Schokolade", "Bier", "Alter Betrag", "Bezahlt", "Neuer Betrag"]

# Initialize Database if not already initialized
def initDB():
  with getConn('igmibot') as conn:
    cur = conn.cursor()
    conn.rollback()
    cur.execute("CREATE TABLE IF NOT EXISTS Members(EMail TEXT PRIMARY KEY NOT NULL, FirstName TEXT NOT NULL DEFAULT 'John', LastName TEXT NOT NULL DEFAULT 'Doe');")
    cur.execute("CREATE TABLE IF NOT EXISTS TelegramID(Id BIGINT PRIMARY KEY NOT NULL, EMail TEXT REFERENCES Members(EMail), Message INTEGER, Language CHAR(2) NOT NULL DEFAULT 'en');")
    cur.execute("CREATE TABLE IF NOT EXISTS Bill(Email TEXT PRIMARY KEY NOT NULL, Coffee INTEGER NOT NULL DEFAULT 0, Softdrink INTEGER NOT NULL DEFAULT 0, Water INTEGER NOT NULL DEFAULT 0, Choc INTEGER NOT NULL DEFAULT 0, Beer INTEGER NOT NULL DEFAULT 0, Oldbill INTEGER NOT NULL DEFAULT 0, Paid INTEGER NOT NULL DEFAULT 0);")
    conn.commit()

# Connect a telegram ID with the IGMI Mail
def insertMember(email, lang, tid, message):
  with getConn('igmibot') as conn:
    cur = conn.cursor()
    cur.execute("INSERT INTO TelegramID(Id, EMail, Message, Language) VALUES(%s, %s, %s, %s);", (tid, email, message, lang))
    conn.commit()

# Check if the entered email-adress is registered in the database
def isIGMIMember(email):
  with getConn('igmibot') as conn:
    cur = conn.cursor()
    cur.execute("SELECT EMail FROM Members WHERE Email = %s;", (email,))
    if cur.fetchone():
      return True
    return False

# Check if Telegram ID or email-adress is already connected in the TelegramID Table
def isTGMember(tid = 0, email = ""):
  with getConn('igmibot') as conn:
    cur = conn.cursor()
    cur.execute("SELECT * FROM TelegramID WHERE Id = %s OR Email = %s;", (tid,email))
    if cur.fetchone():
      return True
    return False

# Get an IGMI Member via his email
def getMemberPerMail(email):
  with getConn('igmibot') as conn:
    cur = conn.cursor()
    cur.execute("SELECT * FROM Members WHERE EMail = %s;", (email,))
    return cur.fetchone()

# Get an IGMI Member via the Telegram ID
def getMemberPerID(tid):
  with getConn('igmibot') as conn:
    cur = conn.cursor()
    cur.execute("SELECT EMail FROM TelegramID WHERE Id = %s;", (tid,))
    return getMemberPerMail(cur.fetchone()[0])

# Get the data from a specific person from the TelegramID Table
def getTelegramID(id = 0, mail = ""):
  with getConn('igmibot') as conn:
    cur = conn.cursor()
    cur.execute("SELECT * FROM TelegramID WHERE Id = %s OR Email = %s;", (id, mail))
    return cur.fetchone()

# Get every connected Telegram ID
def getTelegramIDList():
  with getConn('igmibot') as conn:
    cur = conn.cursor()
    cur.execute("SELECT Id FROM TelegramID;")
    return cur.fetchall()

# Get the Bill datas from a specific member
def getBill(tid):
  with getConn('igmibot') as conn:
    cur = conn.cursor()
    cur.execute("SELECT * FROM Bill WHERE Email = (SELECT Email FROM TelegramID WHERE Id = %s);", (tid,))
    return cur.fetchone()

# When a new message had to be sent, the message-id is stored here
def updateMessage(id, message):
  with getConn('igmibot') as conn:
    cur = conn.cursor()
    cur.execute("UPDATE TelegramID SET Message = %s WHERE Id = %s;", (message, id))
    conn.commit()

# When the chosen language changed, the data in the database is updated
def updateLanguage(id, lang):
  with getConn('igmibot') as conn:
    cur = conn.cursor()
    cur.execute("UPDATE TelegramID SET Language = %s WHERE Id = %s;", (lang, id))
    conn.commit()

# Updating the bill data when someone bought something
def updateBill(tid, coffee = 0, soft = 0, water = 0, choc = 0, beer = 0):
  with getConn('igmibot') as conn:
    cur = conn.cursor()
    cur.execute("UPDATE Bill SET Coffee = Coffee+%s, Softdrink = Softdrink+%s, Water = Water+%s, Choc = Choc+%s, Beer = Beer+%s WHERE Email = (SELECT Email FROM TelegramID WHERE Id = %s);", (coffee, soft, water, choc, beer, tid))
    conn.commit()

# Create Excel sheets for IGMI
# Called every 1st of the month at 12:00
def createExcel(excelDir):
  today = date.today()
  thisYear, thisMonth = [today.year, today.month]
  # While we have the first of the next month, we have to reduce it by one for our calculations
  if thisMonth == 1:
    thisYear -= 1
    thisMonth = 12
  else:
    thisMonth -= 1
  # Try to load the recent workbook. If that fails, create a new one
  # wb = workbook, ws = worksheet
  try:
    wb = load_workbook("{0}/{1}.xlsx".format(excelDir, thisYear))
    ws = wb.create_sheet(months[thisMonth-1])
  except:
    wb = Workbook()
    ws = wb.active
    ws.title = months[thisMonth-1]
  ws.append(heads)
  ws.append([""])
  # Fetch datas from the Bill and Members table to fill the worksheet
  with getConn('igmibot') as conn:
    cur = conn.cursor()
    cur.execute("SELECT M.firstname, M.lastname, M.Email, B.Coffee, B.Softdrink, B.Water, B.Choc, B.Beer, B.Oldbill/100, B.Paid/100 FROM Members M, Bill B WHERE M.email = B.email ORDER BY M.lastname, M.firstname, M.email ASC;")
    # Processing one entry at a time to keep RAM clean
    iterator = cur.fetchone()
    while iterator:
      ws.append(list(iterator) + ["=(D{0}*30 + E{0}*100 + F{0}*60 + G{0}*25 + H{0}*100 + I{0}*100 - J{0}*100)/100".format(ws.max_row+1)])
      iterator = cur.fetchone()
  # Set the column width to proper size
  for column in ws.columns:
    # The middle ones (Schokolade, Softgetränke etc.) get a fixed size
    if column[0].value in heads[3:8]:
      ws.column_dimensions[column[0].column].width = 13
    # The other's sizes get calculated
    else:
      length = max(len(str(cell.value)) for cell in column_cells)
      ws.column_dimensions[column_cells[0].column].width = length
      #ws.column_dimensions[column[0].column].bestFit = True
  wb.save("{0}/{1}.xlsx".format(excelDir, thisYear))
  calculateBill()

# Function for end of the month. Summarizes the bought wares and calculates the price.
# Then sets the bought wares to zero again. All prices in cent.
def calculateBill():
  with getConn('igmibot') as conn:
    cur = conn.cursor()
    cur.execute("UPDATE Bill SET Oldbill = (Oldbill + Coffee*30 + Softdrink*100 + Water*60 + Choc*25 + Beer*100 - Paid), Coffee = 0, Softdrink = 0, Water = 0, Choc = 0, Beer = 0, Paid = 0;")
    conn.commit()

# Removes the Email-Telegram Connection
def removeTelegramID(id = 0, mail = ""):
  with getConn('igmibot') as conn:
    cur = conn.cursor()
    cur.execute("DELETE FROM TelegramID WHERE Id = %s OR EMail = %s;", (id, mail))
    conn.commit()
